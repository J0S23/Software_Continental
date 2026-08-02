# Importacion masiva desde Excel (seccion 23: "importar lecturas/clientes/
# equipos desde Excel"). Reutiliza la capa generica de datos (catalogo_modelos
# + servicios_datos.crear_registro), asi que cada fila creada queda registrada
# en Historial exactamente igual que si se hubiera creado una por una desde
# el formulario.
#
# A proposito NO reutiliza servicios_datos.normalizar_payload tal cual: ese
# validador es estricto (bloquea la fila si falta un campo 'requerido'), pero
# en una carga masiva es comun que el usuario solo tenga a mano parte de los
# datos y quiera completarlos despues. _normalizar_fila_importacion() abajo
# es una version relajada especifica para este router: nunca bloquea una fila
# por un campo faltante, sin importar si el catalogo lo marca como requerido.
import inspect
import io
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from catalogo_modelos import obtener_configuracion_tipo
from Modulos.Permisos import verificar_permiso_escritura
from routers.auth import get_current_user
from servicios_datos import crear_registro, obtener_campos, obtener_enumeraciones, obtener_modelo

router = APIRouter()

# Restringido a los 3 tipos que pide la seccion 23 del documento (los que mas
# volumen de carga manual tienen). No se generaliza a "cualquier tipo del
# catalogo" a proposito: entidades como usuarios o facturacion no deberian
# poblarse por lote sin mas control.
TIPOS_IMPORTABLES = {"lecturas", "clientes", "equipos"}

# Restringido a los 3 tipos que mas volumen de carga manual tienen. No se generaliza a "cualquier tipo del
# catalogo" a proposito: entidades como usuarios o facturacion no deberian poblarse por lote sin mas control.
TIPOS_CAMPO_TEXTO_LIBRE = {"text", "select"}

ESTILO_ENCABEZADO_FONT = Font(bold=True, color="FFFFFF")
ESTILO_ENCABEZADO_FILL = PatternFill(start_color="263445", end_color="263445", fill_type="solid")

PLACEHOLDER_FALTANTE = "N.A"


def _validar_tipo_importable(tipo):
    if tipo not in TIPOS_IMPORTABLES:
        opciones = ", ".join(sorted(TIPOS_IMPORTABLES))
        raise HTTPException(
            status_code=400,
            detail=f"Tipo no importable. Usa uno de: {opciones}",
        )
    return obtener_configuracion_tipo(tipo)


def _valor_desde_excel(configuracion_campo, valor):
    #Convierte lo que openpyxl trae de una celda a un formato uniforme (string/numero/bool crudo) antes de pasarlo por
    #_normalizar_fila_importacion. Sin esto, una celda de fecha (que openpyxl entrega como datetime, no como texto 'YYYY-MM-DD')
    #o un booleano nativo de Excel siempre fallarian la validacion aunque el dato sea correcto.
    
    if valor is None or valor == "":
        return None

    tipo_campo = configuracion_campo.get("tipo")

    if tipo_campo == "date" and isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m-%d")

    if tipo_campo == "boolean" and isinstance(valor, bool):
        return "Si" if valor else "No"

    if isinstance(valor, str):
        return valor.strip()

    return valor


def _normalizar_fila_importacion(configuracion, datos):
    #Version relajada de servicios_datos.normalizar_payload, solo para importacion. Diferencias con la version estricta del CRUD manual:
      #- Un campo 'requerido' que falte NUNCA bloquea la fila.
      #- Si el campo faltante es de texto libre (o un 'select' sin Enum de Python detras, ej. prioridad/estado_lectura), se guarda "N.A" para
      #  que quede visible en la tabla que se dejo pendiente.
      #- Si el campo faltante es numero/fecha/booleano/select-con-Enum, se omite: el modelo aplica su propio default o queda NULL (no se puede
      #  guardar "N.A" en una columna Float/DateTime/Enum sin romper el tipo).
    #Los campos que SI vienen con valor se validan/convierten exactamente gual que en normalizar_payload (mismo casteo de enum/numero/booleano/fecha),
    #asi que un dato mal escrito sigue reportandose como error de fila -- esto solo relaja lo FALTANTE, no lo INVALIDO.
    
    valores = {}
    enumeraciones = obtener_enumeraciones(configuracion)

    for configuracion_campo in obtener_campos(configuracion):
        nombre_campo = configuracion_campo["nombre"]
        valor = datos.get(nombre_campo)
        tipo_campo = configuracion_campo.get("tipo", "text")
        tipo_enum = enumeraciones.get(nombre_campo)

        if valor in (None, ""):
            es_texto_libre = tipo_campo in TIPOS_CAMPO_TEXTO_LIBRE and not tipo_enum
            if es_texto_libre:
                valores[nombre_campo] = PLACEHOLDER_FALTANTE
            # numero / fecha / booleano / select-con-enum faltante: no se
            # envia, toma el default (o NULL) que ya tiene esa columna.
            continue

        if tipo_enum:
            try:
                valor = tipo_enum(valor)
            except ValueError as exc:
                opciones = ", ".join(item.value for item in tipo_enum)
                raise ValueError(f"Valor invalido para '{nombre_campo}'. Usa: {opciones}") from exc

        if tipo_campo == "number":
            try:
                valor = float(valor)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"El campo '{nombre_campo}' debe ser numerico") from exc

        if tipo_campo == "boolean":
            valor_normalizado = str(valor).strip().lower()
            if valor_normalizado in ("si", "sí", "true", "1"):
                valor = True
            elif valor_normalizado in ("no", "false", "0"):
                valor = False
            else:
                raise ValueError(f"El campo '{nombre_campo}' debe ser 'Si' o 'No'")

        if tipo_campo == "date":
            try:
                valor = datetime.strptime(valor, "%Y-%m-%d")
            except (TypeError, ValueError) as exc:
                raise ValueError(f"El campo '{nombre_campo}' debe tener formato de fecha YYYY-MM-DD") from exc

        valores[nombre_campo] = valor

    return valores


def _completar_argumentos_repositorio(modelo, valores):
    #_normalizar_fila_importacion() omite a proposito los campos numero/fecha/booleano/select-con-enum que faltan (para no forzar un tipo
    #invalido con 'N.A'). Pero si el repositorio.agregar() de ese tipo exige ese parametro SIN un valor por defecto propio
    #(ej.ClientesRepositorio.agregar(tipo_cliente, estado_cliente, ...) sin default), omitirlo del todo rompe la llamada con un 
    #TypeError de Python, no con un error de validacion legible.

    #Se resuelve de forma generica con inspect en vez de listar caso por caso que campo es requerido en que repositorio: cualquier parametro de
    #agregar() sin valor por defecto que no haya llegado en 'valores' se completa con None. Los parametros que SI tienen su propio default en el
    #repositorio (ej. precio=0) se dejan intactos, para que se use ese default en vez de forzar None.

    if not hasattr(modelo, "agregar"):
        # Fallback ORM directo (servicios_datos.crear_registro usa
        # modelo(**valores) en ese caso): SQLAlchemy no exige todos los
        # kwargs, asi que no hace falta completar nada aqui.
        return valores

    firma = inspect.signature(modelo.agregar)
    for nombre_parametro, parametro in firma.parameters.items():
        si_es_requerido_sin_default = (
            parametro.default is inspect.Parameter.empty
            and parametro.kind in (inspect.Parameter.POSITIONAL_OR_KEYWORD, inspect.Parameter.KEYWORD_ONLY)
        )
        if si_es_requerido_sin_default and nombre_parametro not in valores:
            valores[nombre_parametro] = None
    return valores


@router.get("/api/importar/{tipo}/plantilla")
async def descargar_plantilla_importacion(tipo: str, usuario=Depends(get_current_user)):
    #Excel con solo la fila de encabezados (los nombres de campo exactos que espera el importador), para que el usuario no tenga que
    #adivinar como se llama cada columna. Las columnas se pueden dejar en blanco: ver_normalizar_fila_importacion para que pasa con cada tipo de campo.
    
    configuracion = _validar_tipo_importable(tipo)
    campos = obtener_campos(configuracion)

    wb = Workbook()
    hoja = wb.active
    hoja.title = tipo[:31]
    encabezados = [c["nombre"] for c in campos]
    hoja.append(encabezados)
    for celda in hoja[1]:
        celda.font = ESTILO_ENCABEZADO_FONT
        celda.fill = ESTILO_ENCABEZADO_FILL
    for indice, campo_config in enumerate(campos, start=1):
        hoja.cell(row=2, column=indice, value=f"({campo_config['etiqueta']})")
        hoja.column_dimensions[hoja.cell(row=1, column=indice).column_letter].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="plantilla_{tipo}.xlsx"'},
    )


@router.post("/api/importar/{tipo}")
async def importar_desde_excel(tipo: str, archivo: UploadFile = File(...), usuario=Depends(get_current_user)):
    #Crea un registro por cada fila con datos del Excel. No detiene el proceso ante una fila invalida: la reporta en 'errores'
    #y sigue con las demas. Fila vacia (todas las celdas en blanco) se ignora en silencio.
    #Los campos faltantes (no invalidos, solo ausentes) NO cuentan como error: se completan con "N.A" o se dejan en su default, ver _normalizar_fila_importacion.
    configuracion = _validar_tipo_importable(tipo)
    verificar_permiso_escritura(tipo, usuario)
    modelo = obtener_modelo(configuracion)
    campos = obtener_campos(configuracion)
    campos_por_nombre = {c["nombre"]: c for c in campos}

    contenido = await archivo.read()
    try:
        wb = load_workbook(io.BytesIO(contenido), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo Excel: {exc}") from exc

    hoja = wb.active
    filas = list(hoja.iter_rows(values_only=True))
    if not filas:
        raise HTTPException(status_code=400, detail="El archivo esta vacio")

    encabezados = [str(celda).strip() if celda is not None else "" for celda in filas[0]]
    encabezados_reconocidos = [e for e in encabezados if e in campos_por_nombre]
    if not encabezados_reconocidos:
        raise HTTPException(
            status_code=400,
            detail=(
                "Ninguna columna del Excel coincide con los campos esperados. "
                f"Descarga la plantilla en GET /api/importar/{tipo}/plantilla."
            ),
        )

    creados = 0
    errores = []

    for numero_fila, fila in enumerate(filas[1:], start=2):
        datos_fila_crudos = dict(zip(encabezados, fila))
        # se descartan columnas del excel que no correspondan a un campo del catalogo
        datos_fila = {
            nombre_campo: _valor_desde_excel(campos_por_nombre[nombre_campo], valor)
            for nombre_campo, valor in datos_fila_crudos.items()
            if nombre_campo in campos_por_nombre
        }

        if not any(v not in (None, "") for v in datos_fila.values()):
            continue  # fila completamente vacia, no cuenta como error

        try:
            valores = _normalizar_fila_importacion(configuracion, datos_fila)
            valores = _completar_argumentos_repositorio(modelo, valores)
            crear_registro(modelo, valores, usuario_id=usuario.id, tipo_entidad=tipo)
            creados += 1
        except (ValueError, HTTPException) as exc:
            mensaje = exc.detail if isinstance(exc, HTTPException) else str(exc)
            errores.append({"fila": numero_fila, "error": mensaje})
        except Exception as exc:
            errores.append({"fila": numero_fila, "error": str(exc)})

    return {
        "success": True,
        "message": f"{creados} registro(s) importado(s), {len(errores)} con error(es)",
        "creados": creados,
        "errores": errores,
    }