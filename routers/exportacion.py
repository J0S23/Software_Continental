#DECISION DE DISEÑO: Se separan dos tipos de exportacion en este mismo router:
#  1. Informes calculados (general/cliente/equipo/cartera/tecnico): PDF con
#     reportlab (ya existia) + Excel con openpyxl (nuevo). Cada uno arma su
#     propio documento porque la forma de cada informe es distinta
#     (indicador/valor vs listas de sub-registros).
#  2. Exportador GENERICO de bases completas (seccion 23: "descargar base de
#     clientes/contratos/lecturas/rentabilidad"): reutiliza la capa de datos
#     generica (catalogo_modelos + servicios_datos) para volcar CUALQUIER
#     tipo del catalogo a Excel sin logica especifica por modulo -- exactamente
#     la idea que ya estaba anotada en este archivo antes de implementarla.

import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, ListFlowable, ListItem,
)

from catalogo_modelos import obtener_configuracion_tipo
from routers.auth import get_current_user
from servicios_datos import listar_registros, obtener_campos, obtener_modelo, serializar
from Servicio.Informes_mensuales import (
    informe_general,
    informe_por_cliente,
    informe_por_equipo,
    informe_cartera,
    informe_tecnico,
)

router = APIRouter()


def _moneda(valor):
    if valor is None:
        return "N/D"
    return f"$ {valor:,.0f}".replace(",", ".")


def _porcentaje(valor):
    return f"{valor:.1f}%" if valor is not None else "N/D"


# ============================= PDF: informes =============================

def _tabla_indicadores(filas):
    tabla = Table(filas, colWidths=[9 * cm, 6 * cm])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#263445")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d8dee6")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    return tabla


def _tabla_lista(encabezados, filas):
    """Tabla generica para PDFs con varias filas de datos (cartera/tecnico),
    a diferencia de _tabla_indicadores que es siempre 'indicador: valor'."""
    tabla = Table([encabezados] + filas, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#263445")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d8dee6")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return tabla


def _doc_base():
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=2 * cm, bottomMargin=2 * cm)
    return buffer, doc


def _pdf_informe_general(periodo, datos):
    buffer, doc = _doc_base()
    estilos = getSampleStyleSheet()
    story = [
        Paragraph("Informe Mensual General - Continental", estilos["Title"]),
        Paragraph(f"Periodo: {periodo}", estilos["Normal"]),
        Paragraph(f"Generado: {datetime.utcnow().strftime('%d/%m/%Y %H:%M')} UTC", estilos["Normal"]),
        Spacer(1, 16),
    ]

    filas = [
        ["Indicador", "Valor"],
        ["Contratos activos", datos["total_contratos"]],
        ["Clientes", datos["total_clientes"]],
        ["Equipos", datos["total_equipos"]],
        ["Facturado del mes", _moneda(datos["facturado_mes"])],
        ["Recaudado del mes", _moneda(datos["recaudado_mes"])],
        ["Cartera pendiente", _moneda(datos["cartera_pendiente"])],
        ["Costos del mes", _moneda(datos["costos_mes"])],
        ["Utilidad bruta", _moneda(datos["utilidad_bruta"])],
        ["Margen promedio", _porcentaje(datos["margen_promedio"])],
        ["Toneres entregados", datos["toneres_entregados"] if datos["toneres_entregados"] is not None else "N/D"],
        ["Preventivos del mes", datos["preventivos_del_mes"] if datos["preventivos_del_mes"] is not None else "N/D"],
        ["Correctivos del mes", datos["correctivos_del_mes"] if datos["correctivos_del_mes"] is not None else "N/D"],
        ["Contratos por vencer (60 dias)", datos["contratos_por_vencer"]],
        ["Clientes en mora", datos["clientes_en_mora"]],
        ["Equipos con fallas recurrentes", datos["equipos_fallas_recurrentes"]],
    ]
    story.append(_tabla_indicadores(filas))
    story.append(Spacer(1, 20))

    story.append(Paragraph("Recomendaciones", estilos["Heading2"]))
    if datos["recomendaciones"]:
        items = [ListItem(Paragraph(r, estilos["Normal"])) for r in datos["recomendaciones"]]
        story.append(ListFlowable(items, bulletType="bullet"))
    else:
        story.append(Paragraph("Sin recomendaciones para este periodo.", estilos["Normal"]))

    doc.build(story)
    buffer.seek(0)
    return buffer


def _pdf_informe_cliente(periodo, cliente_id, datos):
    buffer, doc = _doc_base()
    estilos = getSampleStyleSheet()
    story = [
        Paragraph(f"Informe por Cliente #{cliente_id}", estilos["Title"]),
        Paragraph(f"Periodo: {periodo}", estilos["Normal"]),
        Spacer(1, 16),
    ]

    filas = [
        ["Indicador", "Valor"],
        ["Estado del cliente", datos["estado_general"] or "N/D"],
        ["Contratos activos", datos["contratos_activos"]],
        ["Equipos instalados", datos["equipos_instalados"] if datos["equipos_instalados"] is not None else "N/D"],
        ["Valor mensual contratado", _moneda(datos["valor_mensual_contratado"])],
        ["Consumo del mes", datos["consumo"] if datos["consumo"] is not None else "N/D"],
        ["Facturado", _moneda(datos["facturado"])],
        ["Pagado", _moneda(datos["pagado"])],
        ["Saldo pendiente", _moneda(datos["saldo"])],
        ["Costos", _moneda(datos["costos"])],
        ["Utilidad", _moneda(datos["utilidad"])],
        ["Margen", _porcentaje(datos["margen"])],
    ]
    story.append(_tabla_indicadores(filas))

    doc.build(story)
    buffer.seek(0)
    return buffer


def _pdf_informe_equipo(periodo, equipo_id, datos):
    buffer, doc = _doc_base()
    estilos = getSampleStyleSheet()
    story = [
        Paragraph(f"Informe por Equipo #{equipo_id}", estilos["Title"]),
        Paragraph(f"Periodo: {periodo}", estilos["Normal"]),
        Spacer(1, 16),
    ]

    filas = [
        ["Indicador", "Valor"],
        ["Numero de serie", datos["numero_serie"] or "N/D"],
        ["Cliente actual", datos["cliente_id"] if datos["cliente_id"] is not None else "N/D"],
        ["Contrato actual", datos["contrato_id"] if datos["contrato_id"] is not None else "N/D"],
        ["Contador actual B/N", datos["contador_actual_bn"] if datos["contador_actual_bn"] is not None else "N/D"],
        ["Contador actual color", datos["contador_actual_color"] if datos["contador_actual_color"] is not None else "N/D"],
        ["Consumo mensual B/N", datos["consumo_mensual_bn"] if datos["consumo_mensual_bn"] is not None else "N/D"],
        ["Consumo mensual color", datos["consumo_mensual_color"] if datos["consumo_mensual_color"] is not None else "N/D"],
        ["Costos del periodo", _moneda(datos["costos"])],
        ["Estado tecnico", datos["estado_tecnico"] or "N/D"],
    ]
    story.append(_tabla_indicadores(filas))

    doc.build(story)
    buffer.seek(0)
    return buffer


def _pdf_informe_cartera(periodo, datos):
    buffer, doc = _doc_base()
    estilos = getSampleStyleSheet()
    story = [
        Paragraph("Informe de Cartera - Continental", estilos["Title"]),
        Paragraph(f"Periodo: {periodo}", estilos["Normal"]),
        Spacer(1, 16),
    ]

    encabezados = ["Cliente", "Emitidas", "Pagadas", "Vencidas", "Saldo pendiente", "Dias mora max", "Estado"]
    filas = [
        [
            c["cliente_id"], c["facturas_emitidas"], c["facturas_pagadas"], c["facturas_vencidas"],
            _moneda(c["saldo_pendiente"]), c["dias_mora_max"], c["estado"],
        ]
        for c in datos["clientes"]
    ]

    if filas:
        story.append(_tabla_lista(encabezados, filas))
    else:
        story.append(Paragraph("Sin facturas registradas para este periodo.", estilos["Normal"]))

    doc.build(story)
    buffer.seek(0)
    return buffer


def _pdf_informe_tecnico(periodo, datos):
    buffer, doc = _doc_base()
    estilos = getSampleStyleSheet()
    story = [
        Paragraph("Informe Tecnico - Continental", estilos["Title"]),
        Paragraph(f"Periodo: {periodo}", estilos["Normal"]),
        Spacer(1, 16),
    ]

    filas_indicadores = [
        ["Indicador", "Valor"],
        ["Correctivos del mes", datos["correctivos_del_mes"]],
        ["Tiempo promedio de respuesta (dias)", round(datos["tiempo_promedio_respuesta"], 1) if datos["tiempo_promedio_respuesta"] is not None else "N/D"],
        ["Tiempo promedio de solucion (dias)", round(datos["tiempo_promedio_solucion"], 1) if datos["tiempo_promedio_solucion"] is not None else "N/D"],
    ]
    for estado, cantidad in datos["preventivos_por_estado"].items():
        filas_indicadores.append([f"Preventivos '{estado}'", cantidad])
    story.append(_tabla_indicadores(filas_indicadores))
    story.append(Spacer(1, 20))

    if datos["equipos_con_mas_fallas"]:
        story.append(Paragraph("Equipos con mas fallas", estilos["Heading2"]))
        filas = [[str(e["equipo_id"]), str(e["fallas"])] for e in datos["equipos_con_mas_fallas"]]
        story.append(_tabla_lista(["Equipo", "Fallas"], filas))
        story.append(Spacer(1, 16))

    if datos["repuestos_mas_usados"]:
        story.append(Paragraph("Repuestos mas usados", estilos["Heading2"]))
        filas = [[r["descripcion"], str(r["cantidad"])] for r in datos["repuestos_mas_usados"]]
        story.append(_tabla_lista(["Repuesto", "Cantidad"], filas))

    doc.build(story)
    buffer.seek(0)
    return buffer


@router.get("/api/informes/{periodo}/pdf")
async def exportar_informe_general_pdf(periodo: str, usuario=Depends(get_current_user)):
    try:
        datos = informe_general(periodo)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    buffer = _pdf_informe_general(periodo, datos)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="informe_general_{periodo}.pdf"'},
    )


@router.get("/api/informes/{periodo}/cliente/{cliente_id}/pdf")
async def exportar_informe_cliente_pdf(periodo: str, cliente_id: int, usuario=Depends(get_current_user)):
    try:
        datos = informe_por_cliente(periodo, cliente_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if datos["estado_general"] is None and datos["facturado"] == 0 and datos["costos"] == 0:
        raise HTTPException(status_code=404, detail="Cliente no encontrado o sin datos en el periodo")

    buffer = _pdf_informe_cliente(periodo, cliente_id, datos)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="informe_cliente_{cliente_id}_{periodo}.pdf"'},
    )


@router.get("/api/informes/{periodo}/equipo/{equipo_id}/pdf")
async def exportar_informe_equipo_pdf(periodo: str, equipo_id: int, usuario=Depends(get_current_user)):
    try:
        datos = informe_por_equipo(periodo, equipo_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if datos is None:
        raise HTTPException(status_code=404, detail="Equipo no encontrado")

    buffer = _pdf_informe_equipo(periodo, equipo_id, datos)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="informe_equipo_{equipo_id}_{periodo}.pdf"'},
    )


@router.get("/api/informes/{periodo}/cartera/pdf")
async def exportar_informe_cartera_pdf(periodo: str, usuario=Depends(get_current_user)):
    try:
        datos = informe_cartera(periodo)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    buffer = _pdf_informe_cartera(periodo, datos)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="informe_cartera_{periodo}.pdf"'},
    )


@router.get("/api/informes/{periodo}/tecnico/pdf")
async def exportar_informe_tecnico_pdf(periodo: str, usuario=Depends(get_current_user)):
    try:
        datos = informe_tecnico(periodo)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    buffer = _pdf_informe_tecnico(periodo, datos)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="informe_tecnico_{periodo}.pdf"'},
    )


# ============================ Excel: informes ============================

ESTILO_ENCABEZADO_FONT = Font(bold=True, color="FFFFFF")
ESTILO_ENCABEZADO_FILL = PatternFill(start_color="263445", end_color="263445", fill_type="solid")


def _excel_hoja_indicadores(wb, nombre_hoja, filas):
    """filas: lista de [indicador, valor]. Crea/usa una hoja con encabezado
    'Indicador'/'Valor', igual estructura que _tabla_indicadores en PDF."""
    hoja = wb.active if wb.active.title == "Sheet" else wb.create_sheet(nombre_hoja)
    hoja.title = nombre_hoja
    hoja.append(["Indicador", "Valor"])
    for celda in hoja[1]:
        celda.font = ESTILO_ENCABEZADO_FONT
        celda.fill = ESTILO_ENCABEZADO_FILL
    for fila in filas:
        hoja.append(fila)
    hoja.column_dimensions["A"].width = 32
    hoja.column_dimensions["B"].width = 20
    return hoja


def _excel_hoja_tabla(wb, nombre_hoja, encabezados, filas, primera=False):
    hoja = wb.active if primera and wb.active.title == "Sheet" else wb.create_sheet(nombre_hoja)
    hoja.title = nombre_hoja if primera else hoja.title
    hoja.append(encabezados)
    for celda in hoja[1]:
        celda.font = ESTILO_ENCABEZADO_FONT
        celda.fill = ESTILO_ENCABEZADO_FILL
    for fila in filas:
        hoja.append(fila)
    for columna in hoja.columns:
        letra = columna[0].column_letter
        hoja.column_dimensions[letra].width = 18
    return hoja


def _excel_a_bytes(wb):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _streaming_excel(buffer, nombre_archivo):
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'},
    )


@router.get("/api/informes/{periodo}/excel")
async def exportar_informe_general_excel(periodo: str, usuario=Depends(get_current_user)):
    try:
        datos = informe_general(periodo)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    wb = Workbook()
    filas = [
        ["Contratos activos", datos["total_contratos"]],
        ["Clientes", datos["total_clientes"]],
        ["Equipos", datos["total_equipos"]],
        ["Facturado del mes", datos["facturado_mes"]],
        ["Recaudado del mes", datos["recaudado_mes"]],
        ["Cartera pendiente", datos["cartera_pendiente"]],
        ["Costos del mes", datos["costos_mes"]],
        ["Utilidad bruta", datos["utilidad_bruta"]],
        ["Margen promedio (%)", datos["margen_promedio"]],
        ["Toneres entregados", datos["toneres_entregados"]],
        ["Preventivos del mes", datos["preventivos_del_mes"]],
        ["Correctivos del mes", datos["correctivos_del_mes"]],
        ["Contratos por vencer (60 dias)", datos["contratos_por_vencer"]],
        ["Clientes en mora", datos["clientes_en_mora"]],
        ["Equipos con fallas recurrentes", datos["equipos_fallas_recurrentes"]],
    ]
    _excel_hoja_indicadores(wb, "Informe General", filas)

    if datos["recomendaciones"]:
        hoja = wb.create_sheet("Recomendaciones")
        hoja.append(["Recomendacion"])
        for celda in hoja[1]:
            celda.font = ESTILO_ENCABEZADO_FONT
            celda.fill = ESTILO_ENCABEZADO_FILL
        for r in datos["recomendaciones"]:
            hoja.append([r])
        hoja.column_dimensions["A"].width = 70

    return _streaming_excel(_excel_a_bytes(wb), f"informe_general_{periodo}.xlsx")


@router.get("/api/informes/{periodo}/cliente/{cliente_id}/excel")
async def exportar_informe_cliente_excel(periodo: str, cliente_id: int, usuario=Depends(get_current_user)):
    try:
        datos = informe_por_cliente(periodo, cliente_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if datos["estado_general"] is None and datos["facturado"] == 0 and datos["costos"] == 0:
        raise HTTPException(status_code=404, detail="Cliente no encontrado o sin datos en el periodo")

    wb = Workbook()
    filas = [
        ["Estado del cliente", datos["estado_general"]],
        ["Contratos activos", datos["contratos_activos"]],
        ["Equipos instalados", datos["equipos_instalados"]],
        ["Valor mensual contratado", datos["valor_mensual_contratado"]],
        ["Consumo del mes", datos["consumo"]],
        ["Facturado", datos["facturado"]],
        ["Pagado", datos["pagado"]],
        ["Saldo pendiente", datos["saldo"]],
        ["Costos", datos["costos"]],
        ["Utilidad", datos["utilidad"]],
        ["Margen (%)", datos["margen"]],
    ]
    _excel_hoja_indicadores(wb, f"Cliente {cliente_id}", filas)
    return _streaming_excel(_excel_a_bytes(wb), f"informe_cliente_{cliente_id}_{periodo}.xlsx")


@router.get("/api/informes/{periodo}/equipo/{equipo_id}/excel")
async def exportar_informe_equipo_excel(periodo: str, equipo_id: int, usuario=Depends(get_current_user)):
    try:
        datos = informe_por_equipo(periodo, equipo_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if datos is None:
        raise HTTPException(status_code=404, detail="Equipo no encontrado")

    wb = Workbook()
    filas = [
        ["Numero de serie", datos["numero_serie"]],
        ["Cliente actual", datos["cliente_id"]],
        ["Contrato actual", datos["contrato_id"]],
        ["Contador actual B/N", datos["contador_actual_bn"]],
        ["Contador actual color", datos["contador_actual_color"]],
        ["Consumo mensual B/N", datos["consumo_mensual_bn"]],
        ["Consumo mensual color", datos["consumo_mensual_color"]],
        ["Costos del periodo", datos["costos"]],
        ["Estado tecnico", datos["estado_tecnico"]],
    ]
    _excel_hoja_indicadores(wb, f"Equipo {equipo_id}", filas)
    return _streaming_excel(_excel_a_bytes(wb), f"informe_equipo_{equipo_id}_{periodo}.xlsx")


@router.get("/api/informes/{periodo}/cartera/excel")
async def exportar_informe_cartera_excel(periodo: str, usuario=Depends(get_current_user)):
    try:
        datos = informe_cartera(periodo)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    wb = Workbook()
    encabezados = ["Cliente", "Emitidas", "Pagadas", "Vencidas", "Saldo pendiente", "Dias mora max", "Estado"]
    filas = [
        [
            c["cliente_id"], c["facturas_emitidas"], c["facturas_pagadas"], c["facturas_vencidas"],
            c["saldo_pendiente"], c["dias_mora_max"], c["estado"],
        ]
        for c in datos["clientes"]
    ]
    _excel_hoja_tabla(wb, "Cartera", encabezados, filas, primera=True)
    return _streaming_excel(_excel_a_bytes(wb), f"informe_cartera_{periodo}.xlsx")


@router.get("/api/informes/{periodo}/tecnico/excel")
async def exportar_informe_tecnico_excel(periodo: str, usuario=Depends(get_current_user)):
    try:
        datos = informe_tecnico(periodo)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    wb = Workbook()
    filas_indicadores = [
        ["Correctivos del mes", datos["correctivos_del_mes"]],
        ["Tiempo promedio de respuesta (dias)", datos["tiempo_promedio_respuesta"]],
        ["Tiempo promedio de solucion (dias)", datos["tiempo_promedio_solucion"]],
    ]
    for estado, cantidad in datos["preventivos_por_estado"].items():
        filas_indicadores.append([f"Preventivos '{estado}'", cantidad])
    _excel_hoja_indicadores(wb, "Resumen tecnico", filas_indicadores)

    if datos["equipos_con_mas_fallas"]:
        filas = [[e["equipo_id"], e["fallas"]] for e in datos["equipos_con_mas_fallas"]]
        _excel_hoja_tabla(wb, "Equipos con mas fallas", ["Equipo", "Fallas"], filas)

    if datos["repuestos_mas_usados"]:
        filas = [[r["descripcion"], r["cantidad"]] for r in datos["repuestos_mas_usados"]]
        _excel_hoja_tabla(wb, "Repuestos mas usados", ["Repuesto", "Cantidad"], filas)

    return _streaming_excel(_excel_a_bytes(wb), f"informe_tecnico_{periodo}.xlsx")


# ================= Excel: exportador generico de bases completas =================

@router.get("/api/exportar/{tipo}/excel")
async def exportar_base_completa_excel(tipo: str, usuario=Depends(get_current_user)):
    """Descarga la tabla completa de cualquier tipo del catalogo (seccion 23:
    'descargar base de clientes/contratos/lecturas', y ahora tambien
    rentabilidad, ver catalogo_modelos.py). No requiere logica especifica por
    modulo: usa la misma capa generica que ya usa el CRUD de routers/datos.py."""
    configuracion = obtener_configuracion_tipo(tipo)
    modelo = obtener_modelo(configuracion)
    campos = obtener_campos(configuracion)
    registros = listar_registros(modelo)

    encabezados = ["id"] + [c["nombre"] for c in campos]
    filas = []
    for registro in registros:
        serializado = serializar(registro, campos)
        filas.append([serializado.get(clave) for clave in encabezados])

    wb = Workbook()
    _excel_hoja_tabla(wb, tipo[:31], encabezados, filas, primera=True)  # 31 = limite de Excel para nombres de hoja
    return _streaming_excel(_excel_a_bytes(wb), f"{tipo}.xlsx")